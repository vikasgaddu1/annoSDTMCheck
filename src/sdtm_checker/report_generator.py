from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from pathlib import Path
import logging
from datetime import datetime
from .validation_engine import ValidationResult, ValidationSeverity, ValidationEngine
from .sdtm_reader import SDTMDataset

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates validation reports in various formats."""

    def __init__(self, validation_engine: ValidationEngine):
        """
        Initialize the report generator.

        Args:
            validation_engine: The validation engine instance
        """
        self.validation_engine = validation_engine

        # Define common styles
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.normal_font = Font(size=11)
        self.italic_font = Font(italic=True)

        # Define fills for different severities and headers
        self.header_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.error_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.warning_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.info_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Define borders
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.thin_border = self.border  # Alias for backward compatibility

    def generate_report(
        self,
        results: List[ValidationResult],
        datasets: Dict[str, SDTMDataset],
        output_format: str = 'html'
    ) -> str:
        """
        Generate a validation report.

        Args:
            results: List of validation results
            datasets: Dictionary of SDTM datasets
            output_format: Output format ('html' or 'text')

        Returns:
            Report content as a string
        """
        if output_format == 'html':
            return self._generate_html_report(results, datasets)
        else:
            return self._generate_text_report(results, datasets)

    def _generate_html_report(
        self,
        results: List[ValidationResult],
        datasets: Dict[str, SDTMDataset]
    ) -> str:
        """Generate an HTML report."""
        html = []
        html.append("<!DOCTYPE html>")
        html.append("<html>")
        html.append("<head>")
        html.append("<style>")
        html.append("body { font-family: Arial, sans-serif; margin: 20px; }")
        html.append("h1, h2 { color: #333; }")
        html.append(".error { color: #d32f2f; }")
        html.append(".warning { color: #f57c00; }")
        html.append(".info { color: #1976d2; }")
        html.append(
            "table { border-collapse: collapse; width: 100%; margin: 20px 0; }")
        html.append(
            "th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }")
        html.append("th { background-color: #f5f5f5; }")
        html.append("tr:nth-child(even) { background-color: #f9f9f9; }")
        html.append("</style>")
        html.append("</head>")
        html.append("<body>")

        # Summary section
        html.append("<h1>SDTM Validation Report</h1>")
        html.append("<h2>Summary</h2>")
        html.append("<table>")
        html.append(
            "<tr><th>Total Annotations</th><td>{}</td></tr>".format(len(results)))
        html.append("<tr><th>Errors</th><td class='error'>{}</td></tr>".format(
            sum(1 for r in results if r.severity == 'error')
        ))
        html.append("<tr><th>Warnings</th><td class='warning'>{}</td></tr>".format(
            sum(1 for r in results if r.severity == 'warning')
        ))
        html.append("<tr><th>Info</th><td class='info'>{}</td></tr>".format(
            sum(1 for r in results if r.severity == 'info')
        ))
        html.append("</table>")

        # Detailed results section
        html.append("<h2>Detailed Results</h2>")
        html.append("<table>")
        html.append(
            "<tr><th>Domain</th><th>Variable</th><th>Severity</th><th>Message</th></tr>")

        for result in results:
            html.append("<tr>")
            html.append("<td>{}</td>".format(result.annotation.domain or ''))
            html.append(
                "<td>{}</td>".format(result.annotation.variable_name or ''))
            html.append("<td class='{}'>{}</td>".format(
                result.severity,
                result.severity.capitalize()
            ))
            html.append("<td>{}</td>".format(result.message))
            html.append("</tr>")

        html.append("</table>")
        html.append("</body>")
        html.append("</html>")

        return "\n".join(html)

    def _generate_text_report(
        self,
        results: List[ValidationResult],
        datasets: Dict[str, SDTMDataset]
    ) -> str:
        """Generate a text report."""
        lines = []
        lines.append("SDTM Validation Report")
        lines.append("=" * 50)
        lines.append("")

        # Summary section
        lines.append("Summary")
        lines.append("-" * 20)
        lines.append("Total Annotations: {}".format(len(results)))
        lines.append("Errors: {}".format(
            sum(1 for r in results if r.severity == 'error')))
        lines.append("Warnings: {}".format(
            sum(1 for r in results if r.severity == 'warning')))
        lines.append("Info: {}".format(
            sum(1 for r in results if r.severity == 'info')))
        lines.append("")

        # Detailed results section
        lines.append("Detailed Results")
        lines.append("-" * 20)
        for result in results:
            lines.append("Domain: {}".format(result.annotation.domain or ''))
            lines.append("Variable: {}".format(
                result.annotation.variable_name or ''))
            lines.append("Severity: {}".format(result.severity.capitalize()))
            lines.append("Message: {}".format(result.message))
            lines.append("")

        return "\n".join(lines)

    def _create_dashboard_sheet(self, sheet, results: List[ValidationResult]):
        """Create an executive dashboard summary sheet with integrated metadata, without charts."""
        # Add title
        sheet.merge_cells('A1:L1')
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "SDTM Validation Dashboard"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')

        # Add metadata directly under title
        metadata_row = 2
        sheet.merge_cells(f'I{metadata_row}:L{metadata_row}')
        sheet.cell(row=metadata_row,
                   column=9).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet.cell(row=metadata_row, column=9).font = self.italic_font
        sheet.cell(row=metadata_row, column=9).alignment = Alignment(
            horizontal='right')

        # Calculate summary metrics
        total_issues = len(results)
        error_count = sum(1 for r in results if r.severity ==
                          ValidationSeverity.ERROR)
        warning_count = sum(1 for r in results if r.severity ==
                            ValidationSeverity.WARNING)
        info_count = sum(1 for r in results if r.severity ==
                         ValidationSeverity.INFO)
        unique_domains = len(
            set(r.annotation.domain for r in results if r.annotation.domain))
        unique_variables = len(
            set(r.annotation.variable_name for r in results if r.annotation.variable_name))

        # Count by category
        category_counts = {}
        for result in results:
            category_counts[result.category.value] = category_counts.get(
                result.category.value, 0) + 1

        # Count by domain
        domain_counts = {}
        for result in results:
            if result.annotation.domain:
                domain_counts[result.annotation.domain] = domain_counts.get(
                    result.annotation.domain, 0) + 1

        # Add key metrics section - left side
        sheet.merge_cells('A3:E3')
        metrics_title = sheet.cell(row=3, column=1)
        metrics_title.value = "Key Metrics"
        metrics_title.font = Font(size=14, bold=True)

        # Create metrics grid - 2x3 layout (more compact)
        metrics = [
            ("Total Issues", total_issues),
            ("Errors", error_count),
            ("Warnings", warning_count),
            ("Info", info_count),
            ("Unique Domains", unique_domains),
            ("Unique Variables", unique_variables)
        ]

        # Add metrics with styled boxes in 2 columns
        for idx, (label, value) in enumerate(metrics):
            col = (idx % 2) * 2 + 1
            row = 5 + (idx // 2) * 3

            # Label cell
            label_cell = sheet.cell(row=row, column=col)
            label_cell.value = label
            label_cell.font = Font(bold=True)

            # Value cell
            value_cell = sheet.cell(row=row+1, column=col)
            value_cell.value = value
            value_cell.font = Font(size=14, bold=True)

            # Apply styling based on metric type
            if "Error" in label and isinstance(value, int) and value > 0:
                value_cell.font = Font(size=14, bold=True, color="FF0000")
            elif "Warning" in label and isinstance(value, int) and value > 0:
                value_cell.font = Font(size=14, bold=True, color="FFA500")

        # Add severity distribution section - right side
        severity_row = 5
        sheet.merge_cells(f'G{severity_row}:K{severity_row}')
        severity_title = sheet.cell(row=severity_row, column=7)
        severity_title.value = "Distribution by Severity"
        severity_title.font = Font(size=12, bold=True)

        # Headers for severity table
        headers = ["Severity", "Count", "% of Total"]
        for col, header in enumerate(headers, 7):
            cell = sheet.cell(row=severity_row+2, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        # Add severity data
        severity_counts = {
            "ERROR": error_count,
            "WARNING": warning_count,
            "INFO": info_count
        }

        for row, (severity, count) in enumerate(severity_counts.items(), severity_row+3):
            percentage = (count / total_issues) * \
                100 if total_issues > 0 else 0

            # Apply color based on severity
            fill = self._get_severity_fill(ValidationSeverity(severity))

            sheet.cell(row=row, column=7, value=severity).border = self.border
            sheet.cell(row=row, column=7).fill = fill
            sheet.cell(row=row, column=8, value=count).border = self.border
            sheet.cell(row=row, column=8).fill = fill
            sheet.cell(row=row, column=9,
                       value=f"{percentage:.1f}%").border = self.border
            sheet.cell(row=row, column=9).fill = fill

        # Add critical issues section - move to below metrics
        critical_row = 18
        sheet.merge_cells(f'A{critical_row}:E{critical_row}')
        critical_title = sheet.cell(row=critical_row, column=1)
        critical_title.value = "Most Common Issues"
        critical_title.font = Font(size=14, bold=True)

        # Headers for issues table
        headers = ["Category", "Count", "% of Total"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=critical_row+2, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        # Sort categories by count
        sorted_categories = sorted(
            category_counts.items(), key=lambda x: x[1], reverse=True)

        # Add top categories
        for row, (category, count) in enumerate(sorted_categories[:5], critical_row+3):
            percentage = (count / total_issues) * \
                100 if total_issues > 0 else 0

            sheet.cell(row=row, column=1, value=category).border = self.border
            sheet.cell(row=row, column=2, value=count).border = self.border
            sheet.cell(row=row, column=3,
                       value=f"{percentage:.1f}%").border = self.border

        # Add domain distribution section - right side
        if domain_counts:
            domain_row = 18
            sheet.merge_cells(f'G{domain_row}:K{domain_row}')
            domain_title = sheet.cell(row=domain_row, column=7)
            domain_title.value = "Top Domains with Issues"
            domain_title.font = Font(size=14, bold=True)

            # Headers for domain table
            headers = ["Domain", "Issues", "% of Total"]
            for col, header in enumerate(headers, 7):
                cell = sheet.cell(row=domain_row+2, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border

            # Sort domains by count
            sorted_domains = sorted(
                domain_counts.items(), key=lambda x: x[1], reverse=True)

            # Add top domains
            for row, (domain, count) in enumerate(sorted_domains[:5], domain_row+3):
                percentage = (count / total_issues) * \
                    100 if total_issues > 0 else 0

                sheet.cell(row=row, column=7,
                           value=domain).border = self.border
                sheet.cell(row=row, column=8, value=count).border = self.border
                sheet.cell(row=row, column=9,
                           value=f"{percentage:.1f}%").border = self.border

        # Add summary recommendation - at the very bottom
        recommendation_row = 28
        sheet.merge_cells(f'A{recommendation_row}:L{recommendation_row}')
        rec_title = sheet.cell(row=recommendation_row, column=1)
        rec_title.value = "Summary"
        rec_title.font = Font(size=14, bold=True)

        sheet.merge_cells(f'A{recommendation_row+2}:L{recommendation_row+2}')
        summary_cell = sheet.cell(row=recommendation_row+2, column=1)

        if error_count > 0:
            top_error_categories = [cat for cat, _ in sorted_categories if any(
                r.severity == ValidationSeverity.ERROR and r.category.value == cat for r in results)][:2]
            error_domains = [domain for domain, _ in sorted_domains if any(
                r.severity == ValidationSeverity.ERROR and r.annotation.domain == domain for r in results)][:2]

            summary_text = f"Found {error_count} critical errors requiring attention. "
            if top_error_categories:
                summary_text += f"Focus on fixing '{', '.join(top_error_categories)}' issues "
            if error_domains:
                summary_text += f"in the {', '.join(error_domains)} domain(s)."
        else:
            summary_text = "No critical errors found. "
            if warning_count > 0:
                summary_text += f"Review {warning_count} warnings for potential improvements."
            else:
                summary_text += "All validations passed successfully!"

        summary_cell.value = summary_text

    def _create_details_sheet(self, sheet, results: List[ValidationResult]):
        """Create the detailed mismatches sheet."""
        # Add headers with styling
        headers = [
            "Page", "Original Annotation", "Domain", "Variable", "Value", "Category",
            "Severity", "Message", "Suggested Correction", "Timestamp"
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        # Add data with styling
        for row, result in enumerate(results, 2):
            # Add data cells
            sheet.cell(row=row, column=1,
                       value=result.annotation.page_number).border = self.border
            sheet.cell(
                row=row, column=2, value=result.annotation.annotation_text).border = self.border
            sheet.cell(row=row, column=3,
                       value=result.annotation.domain).border = self.border
            sheet.cell(
                row=row, column=4, value=result.annotation.variable_name).border = self.border
            sheet.cell(row=row, column=5,
                       value=result.annotation.value).border = self.border
            sheet.cell(row=row, column=6,
                       value=result.category.value).border = self.border
            sheet.cell(row=row, column=7,
                       value=result.severity.value).border = self.border
            sheet.cell(row=row, column=8,
                       value=result.message).border = self.border
            sheet.cell(row=row, column=9,
                       value=result.suggested_correction).border = self.border
            sheet.cell(row=row, column=10, value=datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S")).border = self.border

            # Apply color based on pattern type or severity
            if not result.annotation.pattern_type:
                # Use light blue for unmatched annotations
                fill = self.header_fill
            else:
                # Use severity-based colors for matched annotations
                fill = self._get_severity_fill(result.severity)

            for col in range(1, 11):
                sheet.cell(row=row, column=col).fill = fill

        # Add filter
        sheet.auto_filter.ref = f"A1:J{len(results) + 1}"

        # Auto-adjust columns
        self._auto_adjust_columns(sheet)

    def _create_overview_sheet(self, sheet, results: List[ValidationResult]):
        """Create the annotations overview sheet."""
        # Add headers with styling
        headers = [
            "Page", "Annotation Text", "Domain", "Label", "Variable",
            "Value", "Validation Status", "Pattern Type"
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        # Add data with styling
        for row, result in enumerate(results, 2):
            # Add data cells
            sheet.cell(row=row, column=1,
                       value=result.annotation.page_number).border = self.border
            sheet.cell(
                row=row, column=2, value=result.annotation.annotation_text).border = self.border
            sheet.cell(row=row, column=3,
                       value=result.annotation.domain).border = self.border
            sheet.cell(row=row, column=4,
                       value=result.annotation.label).border = self.border
            sheet.cell(
                row=row, column=5, value=result.annotation.variable_name).border = self.border
            sheet.cell(row=row, column=6,
                       value=result.annotation.value).border = self.border
            sheet.cell(row=row, column=7,
                       value=result.severity.value).border = self.border
            sheet.cell(row=row, column=8, value=str(
                result.annotation.pattern_type)).border = self.border

            # Apply color based on severity or pattern type
            if not result.annotation.pattern_type:
                # Use light blue for unmatched annotations
                fill = self.header_fill
            else:
                # Use severity-based colors for matched annotations
                fill = self._get_severity_fill(result.severity)

            for col in range(1, 9):
                sheet.cell(row=row, column=col).fill = fill

        # Add filter
        sheet.auto_filter.ref = f"A1:H{len(results) + 1}"

        # Auto-adjust columns
        self._auto_adjust_columns(sheet)

    def _create_unmatched_sheet(self, sheet, results: List[ValidationResult]):
        """Create a sheet for annotations that didn't match any pattern."""
        # Add headers with styling
        headers = [
            "Page", "Annotation Text", "Position", "Timestamp"
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        # Filter for unmatched annotations (those with annotation text but no pattern matched)
        unmatched_results = [
            r for r in results if r.annotation.annotation_text and not r.annotation.pattern_type]

        # Add data with styling
        for row, result in enumerate(unmatched_results, 2):
            # Add data cells
            sheet.cell(row=row, column=1,
                       value=result.annotation.page_number).border = self.border
            sheet.cell(
                row=row, column=2, value=result.annotation.annotation_text).border = self.border
            position_str = f"({result.annotation.position[0]}, {result.annotation.position[1]})" if result.annotation.position else ""
            sheet.cell(row=row, column=3,
                       value=position_str).border = self.border
            sheet.cell(row=row, column=4, value=datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S")).border = self.border

            # Apply light blue fill color (same as header)
            for col in range(1, 5):
                sheet.cell(row=row, column=col).fill = self.header_fill

        # Add filter
        if unmatched_results:
            sheet.auto_filter.ref = f"A1:D{len(unmatched_results) + 1}"

        # Add note about configuration
        note_row = len(unmatched_results) + 3
        sheet.merge_cells(f'A{note_row}:D{note_row}')
        note_cell = sheet.cell(row=note_row, column=1)
        note_cell.value = "Note: To handle these annotations, please update the configuration file with appropriate patterns."
        note_cell.font = self.italic_font
        note_cell.alignment = Alignment(horizontal='center')

        # Auto-adjust columns
        self._auto_adjust_columns(sheet)

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _export_csv_files(self, results: List[ValidationResult], output_path: str):
        """Export CSV files for each sheet."""
        base_path = Path(output_path).with_suffix('')

        # Export detailed results
        details_data = []
        for result in results:
            details_data.append({
                'Page': result.annotation.page_number,
                'Domain': result.annotation.domain,
                'Variable': result.annotation.variable_name,
                'Value': result.annotation.value,
                'Category': result.category.value,
                'Severity': result.severity.value,
                'Message': result.message,
                'Suggested Correction': result.suggested_correction,
                'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })

        details_df = pd.DataFrame(details_data)
        details_df.to_csv(f"{base_path}_details.csv", index=False)

        # Export summary statistics
        summary_data = []
        for result in results:
            summary_data.append({
                'Category': result.category.value,
                'Severity': result.severity.value
            })

        summary_df = pd.DataFrame(summary_data)
        summary_df = summary_df.groupby(
            ['Category', 'Severity']).size().reset_index(name='Count')
        summary_df['Percentage'] = (
            summary_df['Count'] / len(results) * 100).round(1)
        summary_df.to_csv(f"{base_path}_summary.csv", index=False)

        logger.info(f"CSV files exported to {base_path}_*.csv")

    def _get_severity_fill(self, severity: ValidationSeverity) -> PatternFill:
        """Get the appropriate fill color for a severity level."""
        if severity == ValidationSeverity.ERROR:
            return self.error_fill
        elif severity == ValidationSeverity.WARNING:
            return self.warning_fill
        else:
            return self.info_fill

    def _create_unused_datasets_sheet(self, sheet, datasets: Dict[str, SDTMDataset]):
        """Create a sheet showing datasets that are never referenced in annotations."""
        # Add title
        sheet.merge_cells('A1:B1')
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "Unused SDTM Datasets"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')

        # Add description
        sheet.merge_cells('A2:B2')
        desc_cell = sheet.cell(row=2, column=1)
        desc_cell.value = "The following datasets are present in the study but not referenced in any annotations:"
        desc_cell.font = self.italic_font
        desc_cell.alignment = Alignment(horizontal='center')

        # Add headers
        headers = ["Dataset", "Possible Reason"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        # Add data
        unused_datasets = self.validation_engine.get_unused_datasets(datasets)
        for row, dataset in enumerate(unused_datasets, 5):
            # Dataset name
            cell = sheet.cell(row=row, column=1)
            cell.value = dataset
            cell.border = self.border
            cell.fill = self.header_fill

            # Reason
            cell = sheet.cell(row=row, column=2)
            cell.value = "May be using vendor/external data or oversight in annotation"
            cell.border = self.border
            cell.fill = self.header_fill

        # Add table
        if unused_datasets:
            table = Table(displayName="UnusedDatasets",
                          ref=f"A4:B{len(unused_datasets) + 4}")
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            sheet.add_table(table)

        # Auto-adjust columns
        self._auto_adjust_columns(sheet)

    def export_to_excel(
        self,
        results: List[ValidationResult],
        datasets: Dict[str, SDTMDataset],
        output_path: str
    ):
        """
        Export validation results to Excel with multiple sheets.

        Args:
            results: List of validation results
            datasets: Dictionary of SDTM datasets
            output_path: Path to save the Excel file
        """
        try:
            # Create a new workbook
            wb = Workbook()

            # Create and populate sheets
            dashboard_sheet = wb.active
            dashboard_sheet.title = "Dashboard"
            self._create_dashboard_sheet(dashboard_sheet, results)
            self._auto_adjust_columns(dashboard_sheet)

            details_sheet = wb.create_sheet("Details")
            self._create_details_sheet(details_sheet, results)
            self._auto_adjust_columns(details_sheet)

            overview_sheet = wb.create_sheet("Overview")
            self._create_overview_sheet(overview_sheet, results)
            self._auto_adjust_columns(overview_sheet)

            unmatched_sheet = wb.create_sheet("Unmatched")
            self._create_unmatched_sheet(unmatched_sheet, results)

            unused_sheet = wb.create_sheet("Unused Datasets")
            self._create_unused_datasets_sheet(unused_sheet, datasets)
            self._auto_adjust_columns(unused_sheet)

            # Save the workbook
            wb.save(output_path)
            logger.info(f"Excel report saved to {output_path}")

        except Exception as e:
            logger.error(f"Failed to export Excel report: {str(e)}")
            raise
