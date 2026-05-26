import re
import openpyxl
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTAnno
import fitz  # PyMuPDF


def extract_data_refined(fdf_file):
    """
    Extracts annotation and page data from an FDF file using a refined method.

    Args:
        fdf_file: Path to the FDF file.

    Returns:
        A list of tuples containing annotations and their corresponding page numbers.
    """
    data = []
    with open(fdf_file, 'r', encoding='ISO-8859-1') as f:
        lines = f.readlines()
    
    current_page = None
    current_annotation = None
    
    for line in lines:
        page_match = re.search(r"/Page (\d+)/", line)
        annotation_match = re.search(r"/Contents\((.*?)\)/", line)
        
        if page_match:
            current_page = page_match.group(1)
        
        if annotation_match:
            current_annotation = annotation_match.group(1)
            if current_page and current_annotation:
                # Remove \n and \r from the annotation text
                current_annotation = current_annotation.replace('\n', '').replace('\r', '')
                data.append((current_annotation, current_page))
                current_annotation = None  # Reset after capturing

    return data

def write_to_excel(data, excel_file):
    """
    Writes extracted data to an Excel file.

    Args:
        data: List of tuples containing annotation and page data.
        excel_file: Path to the Excel file.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "Annotation"
    sheet.cell(row=1, column=2).value = "Page"
    for i, (annotation, page) in enumerate(data):
        sheet.cell(row=i+2, column=1).value = annotation
        sheet.cell(row=i+2, column=2).value = page
    wb.save(excel_file)

def extract_annotations_from_pdf(pdf_path):
    """
    Extract annotations from a PDF file using PyMuPDF.
    
    Args:
        pdf_path (str): Path to the input PDF file
    
    Returns:
        list: List of tuples containing (annotation_text, page_number)
    """
    annotations = []
    doc = fitz.open(pdf_path)
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        for annot in page.annots():
            if annot.type[0] == 0:  # Text annotation
                text = annot.info.get("content", "")
                if text:
                    annotations.append((text, page_num + 1))
    
    doc.close()
    return annotations

# Replace these with your file paths
# fdf_file = "./myf1001/myf1001-Uniques_SDTM_07MAY2024_VG.fdf"
pdf_path = "./crf/acrf_myf1001.pdf"
excel_file = "./output/annotation_myf1001.xlsx"

# data = extract_data_refined(fdf_file)
# write_to_excel(data, excel_file)

annotations = extract_annotations_from_pdf(pdf_path)
print(annotations)

# Write annotations to Excel file
write_to_excel(annotations, excel_file)


print("Data extracted and written to Excel file successfully!")
